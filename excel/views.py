from django.shortcuts import render, redirect
from .forms import XlsxModelForm, EditForm
from .models import Xlsx, XlsxData
import openpyxl
import datetime
from django.contrib import messages

def upload_file_view(request):
    form = XlsxModelForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        form =XlsxModelForm(request.POST, request.FILES)
        obj = Xlsx.objects.get(activated=False)
        excel_file = request.FILES['file_name']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        excel_data = []

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            row_data = []
            for cell in row:
                if cell.value == "Position*":
                    continue
                if cell.value:
                    row_data.append(cell.value)
                else:
                    break
            excel_data.append(row_data)
        obj.activated = True
        obj.save()
        for i in range(len(excel_data)):
            XlsxData.objects.create(
                First_name=excel_data[i][0],
                Last_name = excel_data[i][1],
                Mobile = excel_data[i][2],
                Start_date = datetime.datetime.strptime(excel_data[i][3], "%d/%m/%Y").date() if isinstance(excel_data[i][3], str) else excel_data[i][3].date(),
                Salary = excel_data[i][5],
                Employee_ID = excel_data[i][6],

            )


    contex = {
        'form': form,
        'Xlsx': XlsxData.objects.order_by('-id'),
    }
    return render(request, 'excel/upload.html', contex)


def edit_details(request, pk):
    person = XlsxData.objects.get(pk=pk)
    if request.method == "GET":
        form = EditForm(instance=person)
        context = {
            'form': form,
            'person': person,
        }
        return render(request, 'excel/edit.html', context)
    else:
        form = EditForm(request.POST, instance=person)
        if form.is_valid():
            form.save()
            return redirect('excel:upload-view')
        context = {
            'form': form,
            'person': person,
        }
        return render(request, 'excel/edit.html', context)
